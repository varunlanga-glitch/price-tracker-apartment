"""
Condo Tracker - Excel Report Generator

Produces a multi-sheet workbook with:
  Sheet 1: Active Listings       — full sortable master list
  Sheet 2: Top Deals             — deal score >= 65, auto-filtered
  Sheet 3: Relist Alerts         — same address, new MLS# (disguised price drops)
  Sheet 4: Price Drops Today     — listings with price reduced since yesterday
  Sheet 5: Price History Log     — every recorded price change
  Sheet 6: Market Trends         — daily aggregated stats with summary
  Sheet 7: Buyer's Guide         — column explanations and tips
"""

import os
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.filters import AutoFilter

from database import (
    get_active_listings, get_relist_alerts,
    get_all_price_history, get_market_trends, get_today_price_drops
)
from config import REPORTS_DIR, MAX_BUILDING_AGE_YEARS

# Compute the "old building" threshold once at import time
_CUR_YEAR      = date.today().year
_MIN_YEAR_BUILT = (_CUR_YEAR - MAX_BUILDING_AGE_YEARS) if MAX_BUILDING_AGE_YEARS > 0 else 0


# ── Colour palette ────────────────────────────────────────────────────────────
C = {
    "header_bg":     "1F3864",   # dark navy
    "header_fg":     "FFFFFF",
    "subheader_bg":  "2E75B6",   # medium blue
    "subheader_fg":  "FFFFFF",
    "accent_green":  "E2EFDA",   # light green fill
    "accent_red":    "FCE4D6",   # light red/orange fill
    "accent_yellow": "FFEB9C",   # light yellow
    "accent_blue":   "DDEEFF",
    "dark_green":    "375623",
    "dark_red":      "9C0006",
    "dark_orange":   "C65911",
    "dark_blue":     "203864",
    "neutral_grey":  "F2F2F2",
    "border_grey":   "BFBFBF",
    "score_high":    "63BE7B",   # green end of colour scale
    "score_mid":     "FFEB84",   # yellow
    "score_low":     "F8696B",   # red end
    "white":         "FFFFFF",
    "relist_bg":     "FFF2CC",   # amber for relist rows
    "senior_bg":     "E2EFDA",
}


# ── Style factories ───────────────────────────────────────────────────────────

def _fill(hex_color: str):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, name="Arial"):
    return Font(bold=bold, color=color, size=size, name=name)

def _border(style="thin", color="BFBFBF"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _header_row(ws, row: int, labels: list, col_widths: list = None):
    """Write a styled header row."""
    for col, label in enumerate(labels, 1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font      = _font(bold=True, color=C["header_fg"], size=10)
        cell.fill      = _fill(C["header_bg"])
        cell.alignment = _align("center")
        cell.border    = _border(color=C["header_fg"])
    if col_widths:
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

def _freeze_and_filter(ws, freeze_cell: str, filter_ref: str):
    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = filter_ref


# ── Hyperlink helper ──────────────────────────────────────────────────────────

def _set_hyperlink(cell, url: str, label: str = "View listing"):
    """Make a cell a clickable hyperlink styled like a standard Excel link."""
    if not url:
        cell.value = ""
        return
    cell.value     = label
    cell.hyperlink = url
    cell.font      = Font(name="Arial", size=9, color="0563C1", underline="single")


# ── Number formats ────────────────────────────────────────────────────────────
FMT_DOLLAR   = '"$"#,##0'
FMT_DOLLAR_D = '"$"#,##0.00'
FMT_PCT      = '0.0"%"'
FMT_INT      = '#,##0'
FMT_DEC1     = '#,##0.0'
FMT_DATE     = "YYYY-MM-DD"


# ── Shared: colour-scale on a column range ────────────────────────────────────

def _score_color_scale(ws, col_letter: str, start_row: int, end_row: int):
    """Apply green→yellow→red color scale to deal score column."""
    ref = f"{col_letter}{start_row}:{col_letter}{end_row}"
    rule = ColorScaleRule(
        start_type="num", start_value=0,   start_color=C["score_low"],
        mid_type="num",   mid_value=50,    mid_color=C["score_mid"],
        end_type="num",   end_value=100,   end_color=C["score_high"],
    )
    ws.conditional_formatting.add(ref, rule)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Active Listings
# ═══════════════════════════════════════════════════════════════════════════════

ACTIVE_COLS = [
    ("Address",           30),
    ("City",              12),
    ("MLS #",             12),
    ("Price",             12),
    ("$/sqft",            10),
    ("Sqft",               8),
    ("Beds",               6),
    ("Baths",              6),
    ("Year Built",         10),
    ("DOM\n(this listing)",11),
    ("Eff. DOM\n(all relists)", 12),
    ("Maint. Fee/mo",      13),
    ("Tax/yr",             10),
    ("Original Price",     14),
    ("Price Drop $",       12),
    ("Price Drop %",       12),
    ("Parking",             8),
    ("Deal Score",         10),
    ("Relist?",             8),
    ("First Listed",       12),
    ("Last Seen",          12),
    ("Agent",              18),
    ("Brokerage",          20),
    ("URL",                40),
]

def _write_active_sheet(ws, listings: list[dict], summary: dict, today: str):
    # ── Title row ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:X1")
    title = ws["A1"]
    title.value     = f"Abbotsford / Mission / Langley  |  1-Bedroom Condo Tracker  |  {today}"
    title.font      = _font(bold=True, color=C["white"], size=13)
    title.fill      = _fill(C["header_bg"])
    title.alignment = _align("center")
    ws.row_dimensions[1].height = 24

    # ── Summary stats row ─────────────────────────────────────────────────────
    stats = [
        f"Active listings: {summary.get('total_active', 0)}",
        f"Median price: ${summary.get('median_price', 0):,.0f}",
        f"Median $/sqft: ${summary.get('median_psf', 0):,.0f}",
        f"Avg DOM: {summary.get('avg_dom', 0):.0f} days",
        f"Relists: {summary.get('relist_count', 0)} ({summary.get('relist_pct', 0)}%)",
        f"With price drops: {summary.get('with_price_drops', 0)}",
    ]
    col = 1
    for stat in stats:
        cell = ws.cell(row=2, column=col, value=stat)
        cell.font      = _font(bold=True, color=C["dark_blue"], size=10)
        cell.fill      = _fill(C["accent_blue"])
        cell.alignment = _align("left")
        col += 4
    ws.row_dimensions[2].height = 18

    # ── Column headers ────────────────────────────────────────────────────────
    labels = [c[0] for c in ACTIVE_COLS]
    widths = [c[1] for c in ACTIVE_COLS]
    _header_row(ws, 3, labels, widths)
    ws.row_dimensions[3].height = 30

    # ── Data rows ─────────────────────────────────────────────────────────────
    active = sorted(
        [l for l in listings if l["is_active"] and not l.get("is_senior_flagged")],
        key=lambda x: x.get("deal_score") or 0,
        reverse=True
    )

    for r_idx, lst in enumerate(active, start=4):
        row_data = [
            lst.get("address_raw", "").replace("|", " "),
            lst.get("city", ""),
            lst.get("mls_number", ""),
            lst.get("price") or 0,
            lst.get("price_per_sqft") or 0,
            lst.get("sqft") or 0,
            lst.get("bedrooms") or 0,
            lst.get("bathrooms") or 0,
            lst.get("year_built") or "",
            lst.get("dom") or 0,
            lst.get("effective_dom") or lst.get("dom") or 0,
            lst.get("maintenance_fee") or 0,
            lst.get("taxes_annual") or 0,
            lst.get("price_original") or lst.get("price") or 0,
            max(0, (lst.get("price_original") or lst.get("price") or 0) - (lst.get("price") or 0)),
            lst.get("price_reduction_pct") or 0,
            lst.get("parking_spaces") or 0,
            lst.get("deal_score") or 0,
            "YES" if lst.get("is_relist") else "",
            lst.get("first_seen_date", ""),
            lst.get("last_seen_date", ""),
            lst.get("agent_name", ""),
            lst.get("brokerage", ""),
            lst.get("listing_url", ""),
        ]

        # Number formats per column
        col_formats = [
            None, None, None,
            FMT_DOLLAR, FMT_DOLLAR_D, FMT_INT,
            FMT_INT, FMT_DEC1, FMT_INT,
            FMT_INT, FMT_INT,
            FMT_DOLLAR_D, FMT_DOLLAR,
            FMT_DOLLAR, FMT_DOLLAR, FMT_PCT,
            FMT_INT, FMT_DEC1,
            None, FMT_DATE, FMT_DATE,
            None, None, None,
        ]

        row_fill = _fill(C["relist_bg"]) if lst.get("is_relist") else \
                   _fill(C["white"])     if r_idx % 2 == 0 else \
                   _fill(C["neutral_grey"])

        url_col = len(row_data)   # last column = URL
        for c_idx, (val, fmt) in enumerate(zip(row_data, col_formats), start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.fill      = row_fill
            cell.alignment = _align("left" if c_idx in (1, 22, 23, 24) else "center")
            cell.border    = _border(style="hair")
            if c_idx == url_col:
                _set_hyperlink(cell, val)
            else:
                cell.value = val
                cell.font  = _font(size=9)
                if fmt:
                    cell.number_format = fmt

    # ── Conditional formatting ────────────────────────────────────────────────
    last_row = 3 + len(active)
    if last_row > 3:
        # Deal score color scale (col 18 = R)
        _score_color_scale(ws, "R", 4, last_row)

        # Price drop % — highlight red if > 0 (actual drop)
        ws.conditional_formatting.add(
            f"P4:P{last_row}",
            CellIsRule(operator="greaterThan", formula=["0"],
                       font=_font(color=C["dark_green"]),
                       fill=_fill(C["accent_green"]))
        )

        # Effective DOM > 60 — amber warning
        ws.conditional_formatting.add(
            f"K4:K{last_row}",
            CellIsRule(operator="greaterThan", formula=["60"],
                       fill=_fill(C["accent_yellow"]))
        )

        # Relist YES — bold orange text
        ws.conditional_formatting.add(
            f"S4:S{last_row}",
            CellIsRule(operator="equal", formula=['"YES"'],
                       font=_font(bold=True, color=C["dark_orange"]),
                       fill=_fill(C["relist_bg"]))
        )

        # Old building — highlight Year Built cell in amber when the building
        # predates the threshold set in MAX_BUILDING_AGE_YEARS.
        # Year Built is column I (index 9).  Use a FormulaRule so we can
        # test two conditions: value is known (>0) AND below threshold.
        if _MIN_YEAR_BUILT > 0:
            ws.conditional_formatting.add(
                f"I4:I{last_row}",
                FormulaRule(
                    formula=[f"AND(I4>0,I4<{_MIN_YEAR_BUILT})"],
                    fill=_fill(C["accent_yellow"]),
                    font=_font(color=C["dark_orange"], bold=True),
                )
            )

    _freeze_and_filter(ws, "A4", f"A3:{get_column_letter(len(ACTIVE_COLS))}3")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Top Deals
# ═══════════════════════════════════════════════════════════════════════════════

def _write_top_deals_sheet(ws, listings: list[dict]):
    ws.merge_cells("A1:N1")
    t = ws["A1"]
    t.value = "Top Deals  —  Score ≥ 65  |  Sorted by Deal Score"
    t.font = _font(bold=True, color=C["white"], size=12)
    t.fill = _fill(C["subheader_bg"])
    t.alignment = _align("center")

    cols = [
        ("Deal Score", 10), ("Address", 32), ("City", 12), ("MLS #", 12),
        ("Price", 12), ("$/sqft", 10), ("Sqft", 8), ("Year Built", 10),
        ("Eff. DOM", 9), ("Price Drop %", 12), ("Maint. Fee/mo", 13),
        ("Relist?", 8), ("Agent", 18), ("URL", 35),
    ]
    _header_row(ws, 2, [c[0] for c in cols], [c[1] for c in cols])

    top = sorted(
        [l for l in listings if l["is_active"] and not l.get("is_senior_flagged")
         and (l.get("deal_score") or 0) >= 65],
        key=lambda x: x.get("deal_score") or 0,
        reverse=True
    )

    for r_idx, lst in enumerate(top, start=3):
        row_data = [
            lst.get("deal_score") or 0,
            lst.get("address_raw", "").replace("|", " "),
            lst.get("city", ""),
            lst.get("mls_number", ""),
            lst.get("price") or 0,
            lst.get("price_per_sqft") or 0,
            lst.get("sqft") or 0,
            lst.get("year_built") or "",
            lst.get("effective_dom") or lst.get("dom") or 0,
            lst.get("price_reduction_pct") or 0,
            lst.get("maintenance_fee") or 0,
            "YES" if lst.get("is_relist") else "",
            lst.get("agent_name", ""),
            lst.get("listing_url", ""),
        ]
        fmts = [FMT_DEC1, None, None, None, FMT_DOLLAR, FMT_DOLLAR_D,
                FMT_INT, FMT_INT, FMT_INT, FMT_PCT, FMT_DOLLAR_D,
                None, None, None]

        fill = _fill(C["accent_green"]) if r_idx % 2 == 0 else _fill(C["white"])
        url_col = len(row_data)
        for c_idx, (val, fmt) in enumerate(zip(row_data, fmts), start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.fill = fill
            cell.alignment = _align("left" if c_idx in (2, 13, 14) else "center")
            cell.border = _border(style="hair")
            if c_idx == url_col:
                _set_hyperlink(cell, val)
            else:
                cell.value = val
                cell.font = _font(size=9)
                if fmt:
                    cell.number_format = fmt

    if len(top) > 0:
        _score_color_scale(ws, "A", 3, 2 + len(top))

    _freeze_and_filter(ws, "A3", f"A2:{get_column_letter(len(cols))}2")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Relist Alerts
# ═══════════════════════════════════════════════════════════════════════════════

def _write_relist_sheet(ws, listings: list[dict]):
    ws.merge_cells("A1:N1")
    t = ws["A1"]
    t.value = ("⚠  Relist Alerts  —  Same address, new MLS# = disguised price drop  "
               "|  'Effective DOM' reflects true time on market")
    t.font = _font(bold=True, color=C["white"], size=11)
    t.fill = _fill("C65911")
    t.alignment = _align("center")

    cols = [
        ("Address", 32), ("City", 12),
        ("Current MLS #", 13), ("Original MLS #", 13),
        ("Current Price", 13), ("Original Price", 13),
        ("Real Drop $", 12), ("Real Drop %", 12),
        ("DOM (this listing)", 14), ("Effective DOM (all)", 14),
        ("Year Built", 10), ("Sqft", 8), ("$/sqft", 10),
        ("Current URL", 35),
    ]
    _header_row(ws, 2, [c[0] for c in cols], [c[1] for c in cols])

    relists = [l for l in listings if l.get("is_relist") and
               l.get("is_active") and not l.get("is_senior_flagged")]
    relists.sort(key=lambda x: x.get("price_reduction_pct") or 0, reverse=True)

    for r_idx, lst in enumerate(relists, start=3):
        orig_price = lst.get("price_original") or lst.get("price") or 0
        curr_price = lst.get("price") or 0
        drop_amt   = max(0, orig_price - curr_price)
        drop_pct   = lst.get("price_reduction_pct") or 0

        row_data = [
            lst.get("address_raw", "").replace("|", " "),
            lst.get("city", ""),
            lst.get("mls_number", ""),
            lst.get("prior_mls_number", ""),
            curr_price,
            orig_price,
            drop_amt,
            drop_pct,
            lst.get("dom") or 0,
            lst.get("effective_dom") or lst.get("dom") or 0,
            lst.get("year_built") or "",
            lst.get("sqft") or 0,
            lst.get("price_per_sqft") or 0,
            lst.get("listing_url", ""),
        ]
        fmts = [None, None, None, None,
                FMT_DOLLAR, FMT_DOLLAR, FMT_DOLLAR, FMT_PCT,
                FMT_INT, FMT_INT, FMT_INT, FMT_INT, FMT_DOLLAR_D, None]

        url_col = len(row_data)
        for c_idx, (val, fmt) in enumerate(zip(row_data, fmts), start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.fill = _fill(C["relist_bg"])
            cell.alignment = _align("left" if c_idx in (1, 14) else "center")
            cell.border = _border(style="hair")
            if c_idx == url_col:
                _set_hyperlink(cell, val)
            else:
                cell.value = val
                cell.font = _font(size=9,
                                  bold=(c_idx in (7, 8)),
                                  color=(C["dark_red"] if c_idx == 8 and drop_pct > 0 else "000000"))
                if fmt:
                    cell.number_format = fmt

    _freeze_and_filter(ws, "A3", f"A2:{get_column_letter(len(cols))}2")

    if not relists:
        ws.cell(row=3, column=1).value = "No relist alerts — check back tomorrow as history builds."


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Price Drops Today
# ═══════════════════════════════════════════════════════════════════════════════

def _write_price_drops_sheet(ws, today: str):
    drops = get_today_price_drops(today)

    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = f"Price Drops Detected  —  {today}  |  Comparing to yesterday's recorded price"
    t.font = _font(bold=True, color=C["white"], size=12)
    t.fill = _fill("9C0006")
    t.alignment = _align("center")

    cols = [
        ("Address", 32), ("City", 12), ("MLS #", 12),
        ("Price Yesterday", 14), ("Price Today", 14),
        ("Drop $", 12), ("Drop %", 10),
        ("DOM", 8), ("Deal Score", 10), ("URL", 35),
    ]
    _header_row(ws, 2, [c[0] for c in cols], [c[1] for c in cols])

    for r_idx, d in enumerate(drops, start=3):
        row_data = [
            d.get("address_raw", "").replace("|", " "),
            d.get("city", ""),
            d.get("mls_number", ""),
            d.get("price_yesterday") or 0,
            d.get("price_today") or 0,
            abs(d.get("drop_amount") or 0),
            abs(d.get("drop_pct") or 0),
            d.get("dom") or 0,
            d.get("deal_score") or 0,
            d.get("listing_url", ""),
        ]
        fmts = [None, None, None, FMT_DOLLAR, FMT_DOLLAR,
                FMT_DOLLAR, FMT_PCT, FMT_INT, FMT_DEC1, None]

        url_col = len(row_data)
        for c_idx, (val, fmt) in enumerate(zip(row_data, fmts), start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.fill = _fill(C["accent_red"])
            cell.alignment = _align("left" if c_idx in (1, 10) else "center")
            cell.border = _border(style="hair")
            if c_idx == url_col:
                _set_hyperlink(cell, val)
            else:
                cell.value = val
                cell.font = _font(size=9,
                                  bold=(c_idx in (6, 7)),
                                  color=(C["dark_red"] if c_idx == 7 else "000000"))
                if fmt:
                    cell.number_format = fmt

    if not drops:
        ws.cell(row=3, column=1).value = "No price drops recorded today (or this is the first run)."

    _freeze_and_filter(ws, "A3", f"A2:{get_column_letter(len(cols))}2")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Price History Log
# ═══════════════════════════════════════════════════════════════════════════════

def _write_history_sheet(ws):
    history = get_all_price_history()

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "Price History Log  —  Every daily snapshot recorded per listing"
    t.font = _font(bold=True, color=C["white"], size=12)
    t.fill = _fill(C["header_bg"])
    t.alignment = _align("center")

    cols = [
        ("Date", 12), ("MLS #", 12), ("Address", 32),
        ("City", 12), ("Price", 12), ("DOM at Date", 12), ("Notes", 20),
    ]
    _header_row(ws, 2, [c[0] for c in cols], [c[1] for c in cols])

    # Group by MLS to detect changes
    prev_price = {}
    for r_idx, h in enumerate(history, start=3):
        mls   = h.get("mls_number", "")
        price = h.get("price") or 0
        note  = ""
        if mls in prev_price:
            diff = price - prev_price[mls]
            if diff < 0:
                note = f"↓ Drop ${abs(diff):,.0f}"
            elif diff > 0:
                note = f"↑ Increase ${diff:,.0f}"
        prev_price[mls] = price

        row_data = [
            h.get("snapshot_date", ""),
            mls,
            h.get("address_raw", "").replace("|", " "),
            h.get("city", ""),
            price,
            h.get("dom") or 0,
            note,
        ]
        fmts = [FMT_DATE, None, None, None, FMT_DOLLAR, FMT_INT, None]

        fill = _fill(C["accent_red"]) if "Drop" in note else \
               _fill(C["white"]) if r_idx % 2 == 0 else _fill(C["neutral_grey"])

        for c_idx, (val, fmt) in enumerate(zip(row_data, fmts), start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = _font(size=9, bold=("Drop" in note and c_idx == 7),
                              color=(C["dark_red"] if "Drop" in note and c_idx == 7 else "000000"))
            cell.fill = fill
            cell.alignment = _align("left" if c_idx == 3 else "center")
            cell.border = _border(style="hair")
            if fmt:
                cell.number_format = fmt

    _freeze_and_filter(ws, "A3", f"A2:{get_column_letter(len(cols))}2")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — Market Trends
# ═══════════════════════════════════════════════════════════════════════════════

def _write_trends_sheet(ws):
    trends = get_market_trends()

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "Market Trends  —  Daily Aggregated Stats"
    t.font = _font(bold=True, color=C["white"], size=12)
    t.fill = _fill(C["subheader_bg"])
    t.alignment = _align("center")

    cols = [
        ("Date", 12), ("Active Listings", 14), ("Avg Price", 13),
        ("Min Price", 12), ("Max Price", 12), ("Avg $/sqft", 12),
        ("Avg DOM", 10), ("Relist Count", 12),
    ]
    _header_row(ws, 2, [c[0] for c in cols], [c[1] for c in cols])

    for r_idx, t_row in enumerate(trends, start=3):
        row_data = [
            t_row.get("snapshot_date", ""),
            t_row.get("active_count") or 0,
            t_row.get("avg_price") or 0,
            t_row.get("min_price") or 0,
            t_row.get("max_price") or 0,
            t_row.get("avg_price_per_sqft") or 0,
            t_row.get("avg_dom") or 0,
            t_row.get("relist_count") or 0,
        ]
        fmts = [FMT_DATE, FMT_INT, FMT_DOLLAR, FMT_DOLLAR, FMT_DOLLAR,
                FMT_DOLLAR_D, FMT_DEC1, FMT_INT]

        fill = _fill(C["white"]) if r_idx % 2 == 0 else _fill(C["neutral_grey"])
        for c_idx, (val, fmt) in enumerate(zip(row_data, fmts), start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = _font(size=9)
            cell.fill = fill
            cell.alignment = _align("center")
            cell.border = _border(style="hair")
            if fmt:
                cell.number_format = fmt

    if not trends:
        ws.cell(row=3, column=1).value = "Trends will appear after multiple daily runs."

    _freeze_and_filter(ws, "A3", f"A2:{get_column_letter(len(cols))}2")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — Buyer's Guide
# ═══════════════════════════════════════════════════════════════════════════════

GUIDE_CONTENT = [
    ("FIELD",                 "WHAT IT MEANS",                                                                                                       "BUYER TIP"),
    ("Price",                 "Current asking price in CAD",                                                                                          "Always negotiate — use price history and DOM as leverage"),
    ("$/sqft",                "Price divided by square footage — best apples-to-apples comparison",                                                    "Abbotsford average is typically $400–$600/sqft for newer condos. Anything below avg is worth a look."),
    ("Sqft",                  "Interior living area in square feet",                                                                                   "Ask for strata plan to verify — marketing sqft often includes balconies"),
    ("Year Built",            "Construction year of the building",                                                                                     "This tracker shows 2011+ only. Newer = lower maintenance risk and more modern layouts."),
    ("DOM",                   "Days on market for this specific listing",                                                                              "30-60 days = seller getting anxious. 60+ days = negotiate hard."),
    ("Eff. DOM",              "Total days including any prior listings at same address (relist detection)",                                             "A unit relisted after 45 days = truly 45+ DOM even if the new listing shows '1 day'. This is the real number."),
    ("Maint. Fee/mo",         "Monthly strata fee — covers building insurance, common area maintenance, contingency reserve fund",                      "Ask for strata depreciation report & reserve fund study. Low fees + healthy reserve = good sign. Very low fees may mean underfunding."),
    ("Tax/yr",                "Annual property tax as of listing date",                                                                                "Verify with BC Assessment. May change after purchase."),
    ("Original Price",        "The price when this MLS# (or a prior MLS# at same address) was first listed",                                           "Tracks true price history — even across relists. A $450K listing originally at $499K has already dropped 10%."),
    ("Price Drop $",          "Dollar reduction from original to current price",                                                                        "Use this in negotiations — sellers at a drop are more motivated."),
    ("Price Drop %",          "Percentage reduction from original price",                                                                              "Highlight this to your realtor: 'This unit has already dropped X% — what does that tell us?'"),
    ("Deal Score",            "0–100 composite score: price/sqft vs market, price history, DOM, fees, year built",                                     "70+ = strong value signal. 85+ = exceptional. Not a guarantee — do your own diligence."),
    ("Relist?",               "YES if the same address had a different MLS# that expired. Sellers sometimes relist to reset DOM counter.",              "On realtor.ca, a relisted unit appears fresh. Eff. DOM reveals the truth. This is a key negotiating tool."),
    ("",                      "",                                                                                                                       ""),
    ("REALTOR TRICKS",        "What to watch for",                                                                                                     "Counter-strategy"),
    ("Price 'reductions'",    "Some agents reduce by $1-5K just to trigger a 'price reduced' flag on portals",                                         "Look at $ and % drop — ignore token reductions under 1%"),
    ("Relist as new",         "Expired listing relisted at new MLS# to reset DOM to zero",                                                             "This tracker catches it. Eff. DOM tells the real story."),
    ("Maintenance fee omit.", "Some listings omit fees to make the price look better",                                                                 "Always ask for Form B (strata information certificate) from BC Real Estate Association"),
    ("Marketed sqft",         "Often includes balcony, storage locker or shared space in sq footage",                                                   "Request strata plan for interior sq ft only. Balcony adds value but isn't living space."),
    ("'Price upon request'",  "Agent withholding price to force contact",                                                                              "Contact listing agent directly — this tracker only shows listed prices."),
    ("",                      "",                                                                                                                       ""),
    ("DUE DILIGENCE CHECKLIST","Before making an offer:",                                                                                              ""),
    ("Strata docs",           "Request last 2 years of meeting minutes + depreciation report",                                                         "Look for special assessments, disputes, or deferred maintenance"),
    ("Reserve fund",          "Ask for Form B — shows contingency reserve fund balance",                                                               "Under $5K/unit in reserve = possible special assessment risk"),
    ("Rental restrictions",   "Some strata corps limit rentals",                                                                                       "Check bylaws if you plan to rent"),
    ("Pet restrictions",      "Common in older buildings",                                                                                             "Check strata bylaws"),
    ("Insurance",             "Building must carry adequate coverage",                                                                                 "Ask for proof of strata insurance certificate"),
    ("",                      "",                                                                                                                       ""),
    ("ABOUT THIS TRACKER",    "How it works",                                                                                                          ""),
    ("Data source",           "realtor.ca (MLS listings)",                                                                                            "Run daily to build historical data. More history = better trend signals."),
    ("Relist detection",      "Normalizes addresses and links MLS numbers across time",                                                                 "Works better after 2+ weeks of data. Prior relists before first run are not retroactively captured."),
    ("55+ filter",            "Listings with senior/age-restriction keywords are excluded automatically",                                              "Check the 55+ flag column if you suspect a listing is missing."),
    ("Schedule",              "Run: python run.py  — ideally once per day at same time",                                                              "On Windows: Task Scheduler. On Mac/Linux: cron. See README.txt."),
]

def _write_guide_sheet(ws):
    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = "Buyer's Reference Guide  |  Column Explanations & Realtor Tricks"
    t.font = _font(bold=True, color=C["white"], size=13)
    t.fill = _fill(C["header_bg"])
    t.alignment = _align("center")

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 55

    for r_idx, (field, meaning, tip) in enumerate(GUIDE_CONTENT, start=2):
        is_section = meaning == "What to watch for" or meaning == "Before making an offer:" \
                     or meaning == "How it works" or field == "FIELD" or field == "REALTOR TRICKS" \
                     or field == "DUE DILIGENCE CHECKLIST" or field == "ABOUT THIS TRACKER"

        fill = _fill(C["subheader_bg"]) if is_section else \
               _fill(C["white"]) if r_idx % 2 == 0 else _fill(C["neutral_grey"])
        fg   = C["white"] if is_section else "000000"

        for c_idx, val in enumerate([field, meaning, tip], start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = _font(bold=is_section, color=fg, size=9)
            cell.fill = fill
            cell.alignment = _align("left", wrap=True)
            cell.border = _border(style="hair")
        ws.row_dimensions[r_idx].height = 30 if "\n" in (meaning + tip) else 20


# ═══════════════════════════════════════════════════════════════════════════════
# Main report function
# ═══════════════════════════════════════════════════════════════════════════════

def generate_report(listings: list[dict], summary: dict, today: str = None) -> str:
    if today is None:
        today = date.today().isoformat()

    wb = Workbook()

    # Remove default sheet
    default = wb.active
    wb.remove(default)

    sheet_defs = [
        ("Active Listings",   lambda ws: _write_active_sheet(ws, listings, summary, today)),
        ("Top Deals",         lambda ws: _write_top_deals_sheet(ws, listings)),
        ("Relist Alerts",     lambda ws: _write_relist_sheet(ws, listings)),
        ("Price Drops Today", lambda ws: _write_price_drops_sheet(ws, today)),
        ("Price History Log", lambda ws: _write_history_sheet(ws)),
        ("Market Trends",     lambda ws: _write_trends_sheet(ws)),
        ("Buyer's Guide",     lambda ws: _write_guide_sheet(ws)),
    ]

    for name, writer in sheet_defs:
        ws = wb.create_sheet(title=name)
        ws.sheet_properties.tabColor = {
            "Active Listings":   "1F3864",
            "Top Deals":         "375623",
            "Relist Alerts":     "C65911",
            "Price Drops Today": "9C0006",
            "Price History Log": "4472C4",
            "Market Trends":     "2E75B6",
            "Buyer's Guide":     "7F7F7F",
        }.get(name, "000000")
        writer(ws)

    filename  = f"condo_tracker_{today}.xlsx"
    filepath  = os.path.join(REPORTS_DIR, filename)

    # If the file is open in Excel, Windows locks it — try with a counter suffix
    for attempt in range(10):
        candidate = filepath if attempt == 0 else filepath.replace(
            ".xlsx", f"_{attempt}.xlsx"
        )
        try:
            wb.save(candidate)
            return candidate
        except PermissionError:
            if attempt == 0:
                print(f"  [WARN] {filename} is open — saving with suffix instead.")
    # Last resort: force a unique name using seconds
    import time as _time
    fallback = filepath.replace(".xlsx", f"_{int(_time.time())}.xlsx")
    wb.save(fallback)
    return fallback
